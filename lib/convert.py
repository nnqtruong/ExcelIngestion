"""Step 01: Convert Excel in raw/ to Parquet in clean/ (chunked for large files)."""
import sys
import time
from pathlib import Path
from typing import Callable

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import psutil

from lib.logging_util import monitor_step
from lib.schema import columns_as_list, load_schema

EXCEL_EXTENSIONS = (".xlsx", ".xls")
DEFAULT_CHUNK_SIZE = 50_000


def _format_size(size_bytes: int) -> str:
    """Format bytes as human-readable string."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"


def _format_time(seconds: float) -> str:
    """Format seconds as human-readable string."""
    if seconds < 60:
        return f"{seconds:.1f}s"
    else:
        mins = int(seconds // 60)
        secs = seconds % 60
        return f"{mins}m {secs:.1f}s"


def coerce_all_to_string(df: pd.DataFrame) -> pd.DataFrame:
    """Convert ALL columns to string to ensure consistent schema across chunks.

    This prevents PyArrow schema mismatch errors when the same column is inferred
    as int64 in one chunk but string in another (e.g., taskid with mixed values).
    Step 04 (clean_errors) will cast to the correct types using schema.yaml.
    """
    for col in df.columns:
        # Use vectorized approach to avoid "truth value of Series is ambiguous" error
        # which can occur when cells contain array-like objects or merged cell artifacts
        df[col] = df[col].astype(object).where(df[col].notna(), None)
        df[col] = df[col].apply(lambda x: str(x) if x is not None else None)
        df[col] = df[col].astype("string")
    return df


def _schema_defines_source_system(schema: dict) -> bool:
    """True if schema columns include source_system (employees_master, etc.)."""
    try:
        return any(c.get("name") == "source_system" for c in columns_as_list(schema))
    except Exception:
        return False


def convert_excel_to_parquet(
    excel_path: Path,
    output_path: Path,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    source_system: str | None = None,
) -> tuple[int, list[str], int]:
    """Read Excel in chunks via openpyxl, stream to Parquet. Returns (total_rows, headers, chunk_count).

    Uses read_only=True and iter_rows so 500K+ row files stay under ~200MB RAM.
    If source_system is set, a constant source_system column is appended (employees_master).
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]

    writer = None
    total_rows = 0
    chunk_count = 0
    chunk = []

    def _with_source(df: pd.DataFrame) -> pd.DataFrame:
        if source_system is None:
            return df
        out = df.copy()
        out["source_system"] = source_system
        out["source_system"] = out["source_system"].astype("string")
        return out

    for row in rows:
        chunk.append(row)
        if len(chunk) >= chunk_size:
            df = pd.DataFrame(chunk, columns=headers)
            df = coerce_all_to_string(df)
            df = _with_source(df)
            table = pa.Table.from_pandas(df, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(str(output_path), table.schema)
            writer.write_table(table)
            total_rows += len(chunk)
            chunk_count += 1
            chunk = []

    if chunk:
        df = pd.DataFrame(chunk, columns=headers)
        df = coerce_all_to_string(df)
        df = _with_source(df)
        table = pa.Table.from_pandas(df, preserve_index=False)
        if writer is None:
            writer = pq.ParquetWriter(str(output_path), table.schema)
        writer.write_table(table)
        total_rows += len(chunk)
        chunk_count += 1

    if writer is not None:
        writer.close()
    else:
        # No data rows: write empty Parquet with schema from headers
        empty_cols = list(headers) + (["source_system"] if source_system else [])
        empty_df = pd.DataFrame(columns=empty_cols)
        table = pa.Table.from_pandas(empty_df, preserve_index=False)
        pq.write_table(table, str(output_path))

    wb.close()
    out_headers = list(headers)
    if source_system is not None:
        out_headers.append("source_system")
    return total_rows, out_headers, chunk_count


@monitor_step
def run_convert(
    raw_dir: Path,
    clean_dir: Path,
    on_progress: Callable[[int, int, Path, int, float], None] | None = None,
    verbose: bool = True,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
) -> int:
    """Read all Excel from raw_dir, write Parquet to clean_dir (chunked for .xlsx). Returns number of files converted.

    Args:
        raw_dir: Directory containing Excel files
        clean_dir: Directory to write Parquet files
        on_progress: Callback(current, total, path, rows, elapsed_seconds) after each file
        verbose: If True and no callback, print progress to stdout
        chunk_size: Rows per chunk for .xlsx (ignored for .xls)
    """
    if not raw_dir.is_dir():
        raise FileNotFoundError(f"No raw/ directory at {raw_dir}")
    clean_dir.mkdir(parents=True, exist_ok=True)

    excel_files = []
    for ext in EXCEL_EXTENSIONS:
        excel_files.extend(raw_dir.glob(f"*{ext}"))
    excel_files = sorted(excel_files)

    if not excel_files:
        return 0

    add_source_system = False
    try:
        from lib.paths import SCHEMA_PATH

        if SCHEMA_PATH.exists():
            add_source_system = _schema_defines_source_system(load_schema(SCHEMA_PATH))
    except Exception:
        add_source_system = False

    total = len(excel_files)
    total_size = sum(f.stat().st_size for f in excel_files)
    total_rows = 0
    start_time = time.time()
    proc = psutil.Process()

    if verbose and on_progress is None:
        print(f"Converting {total} Excel file(s) ({_format_size(total_size)})...")
        sys.stdout.flush()

    for i, path in enumerate(excel_files, 1):
        file_start = time.time()
        file_size = path.stat().st_size
        rss_before_mb = proc.memory_info().rss / 1024 / 1024

        if verbose and on_progress is None:
            print(f"  [{i}/{total}] {path.name} ({_format_size(file_size)})... ", end="")
            sys.stdout.flush()

        out_path = clean_dir / f"{path.stem}.parquet"

        stem = path.stem
        src_sys = stem if add_source_system else None

        if path.suffix.lower() == ".xlsx":
            rows, headers, chunk_count = convert_excel_to_parquet(
                path, out_path, chunk_size=chunk_size, source_system=src_sys
            )
        else:
            # .xls: openpyxl doesn't support it, use pandas
            df = pd.read_excel(path, engine="xlrd" if path.suffix.lower() == ".xls" else "openpyxl")
            df = coerce_all_to_string(df)
            if src_sys is not None:
                df["source_system"] = src_sys
                df["source_system"] = df["source_system"].astype("string")
            df.to_parquet(out_path, index=False)
            rows = len(df)
            chunk_count = 1

        file_elapsed = time.time() - file_start
        rss_after_mb = proc.memory_info().rss / 1024 / 1024
        total_rows += rows

        if on_progress is not None:
            on_progress(i, total, path, rows, file_elapsed)
        elif verbose:
            print(
                f"{rows:,} rows, {chunk_count} chunk(s), "
                f"memory {rss_before_mb:.1f} -> {rss_after_mb:.1f} MB in {_format_time(file_elapsed)}"
            )
            sys.stdout.flush()

    total_elapsed = time.time() - start_time
    if verbose and on_progress is None:
        print(f"Done. {total_rows:,} total rows in {_format_time(total_elapsed)}")

    return total
