"""Compare Excel headers across raw files to detect schema drift."""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lib.data_root import get_data_root, get_dataset_path

DEFAULT_DATASET = "tasks"
DEFAULT_ENV = "dev"


def read_excel_headers(filepath: Path) -> list[str]:
    """Read only the header row from an Excel file."""
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    worksheet = workbook.active
    headers: list[str] = []
    for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [
            str(cell).strip() if cell is not None and str(cell).strip() else f"_unnamed_{index}"
            for index, cell in enumerate(row)
        ]
        break
    workbook.close()
    return headers


def normalize_name(name: str) -> str:
    """Normalize a column name to detect likely variants."""
    return "".join(char.lower() for char in name if char.isalnum())


def find_variants(columns: set[str]) -> list[list[str]]:
    """Find column names that only differ in formatting."""
    normalized: dict[str, list[str]] = {}
    for column in columns:
        key = normalize_name(column)
        normalized.setdefault(key, []).append(column)
    groups = [sorted(names) for names in normalized.values() if len(names) > 1]
    return sorted(groups, key=lambda names: (len(names), names))


def _list_excel_files(raw_dir: Path) -> list[Path]:
    """Get all Excel files we ingest from raw/."""
    return sorted([*raw_dir.glob("*.xlsx"), *raw_dir.glob("*.xlsm")], key=lambda p: p.name.lower())


def compare_schemas(raw_dir: Path) -> dict[str, Any]:
    """Compare headers across all Excel files in a directory."""
    files = _list_excel_files(raw_dir)
    if not files:
        return {"error": f"No Excel files found in {raw_dir}"}

    file_headers: dict[str, list[str]] = {}
    for excel_file in files:
        file_headers[excel_file.name] = read_excel_headers(excel_file)

    all_columns: set[str] = set()
    for headers in file_headers.values():
        all_columns.update(headers)

    baseline_name = files[0].name
    intersection = set(file_headers[baseline_name])
    for headers in file_headers.values():
        intersection &= set(headers)

    extra_columns: dict[str, list[str]] = {}
    missing_columns: dict[str, list[str]] = {}
    for column in sorted(all_columns):
        has_column = sorted([filename for filename, headers in file_headers.items() if column in headers])
        missing_column = sorted([filename for filename, headers in file_headers.items() if column not in headers])
        if len(has_column) < len(file_headers):
            extra_columns[column] = has_column
        if missing_column:
            missing_columns[column] = missing_column

    return {
        "directory": str(raw_dir),
        "data_root": str(get_data_root()),
        "file_count": len(files),
        "union_columns": sorted(all_columns),
        "intersection_columns": sorted(intersection),
        "possible_variants": find_variants(all_columns),
        "files": {
            filename: {"columns": headers, "column_count": len(headers)}
            for filename, headers in file_headers.items()
        },
        "extra_columns": extra_columns,
        "missing_columns": missing_columns,
    }


def compare_against_baseline(raw_dir: Path, baseline_name: str) -> dict[str, Any]:
    """Compare all files in raw_dir against a specific baseline file."""
    files = _list_excel_files(raw_dir)
    if not files:
        return {"error": f"No Excel files found in {raw_dir}"}

    files_by_name = {file.name: file for file in files}
    baseline_path = files_by_name.get(baseline_name)
    if baseline_path is None:
        return {
            "error": f"Baseline file not found: {baseline_name}",
            "available_files": sorted(files_by_name.keys()),
        }

    baseline_headers = read_excel_headers(baseline_path)
    baseline_set = set(baseline_headers)

    matches: list[str] = []
    differs: dict[str, dict[str, list[str]]] = {}
    for file in files:
        if file.name == baseline_name:
            continue
        headers = read_excel_headers(file)
        current_set = set(headers)
        extra = sorted(current_set - baseline_set)
        missing = sorted(baseline_set - current_set)
        if not extra and not missing:
            matches.append(file.name)
        else:
            differs[file.name] = {"extra": extra, "missing": missing}

    return {
        "baseline": baseline_name,
        "baseline_column_count": len(baseline_headers),
        "matches": sorted(matches),
        "differs": differs,
    }


def check_against_schema(raw_dir: Path, schema_path: Path) -> dict[str, Any]:
    """Compare raw file headers against schema.yaml expectations."""
    from lib.config import load_dataset_config, schema_body_from_merged_config

    files = _list_excel_files(raw_dir)
    if not files:
        return {"error": f"No Excel files found in {raw_dir}"}
    dataset_root = schema_path.parent.parent
    unified = dataset_root / "dataset.yaml"
    if unified.exists():
        schema = schema_body_from_merged_config(load_dataset_config(dataset_root))
    elif not schema_path.exists():
        return {"error": f"Schema file not found: {schema_path}"}
    else:
        with schema_path.open(encoding="utf-8") as handle:
            schema = yaml.safe_load(handle) or {}

    column_order = schema.get("column_order") or []
    column_aliases = schema.get("column_aliases") or {}

    # Filter out pipeline-generated columns (not expected in source Excel files)
    # These include: _source_file, _ingested_at, source_system, row_id, etc.
    pipeline_generated = {"_source_file", "_ingested_at", "source_system", "row_id"}
    expected_columns = set(col for col in column_order if col not in pipeline_generated)
    alias_map = {str(excel_name): str(canonical) for excel_name, canonical in column_aliases.items()}

    files_result: dict[str, dict[str, Any]] = {}
    all_mapped_columns: set[str] = set()
    all_extra_columns: set[str] = set()
    aliases_used: dict[str, set[str]] = {}

    for file in files:
        original_headers = read_excel_headers(file)
        mapped_headers: list[str] = []
        file_aliases_used: dict[str, str] = {}

        for header in original_headers:
            mapped = alias_map.get(header, header)
            mapped_headers.append(mapped)
            if mapped != header:
                file_aliases_used[header] = mapped
                aliases_used.setdefault(mapped, set()).add(header)

        mapped_set = set(mapped_headers)
        all_mapped_columns.update(mapped_set)

        missing_in_file = sorted(expected_columns - mapped_set)
        extra_in_file = sorted(mapped_set - expected_columns)
        all_extra_columns.update(extra_in_file)

        files_result[file.name] = {
            "column_count": len(original_headers),
            "original_headers": original_headers,
            "mapped_headers": mapped_headers,
            "missing_from_schema_columns": missing_in_file,
            "extra_not_in_schema_columns": extra_in_file,
            "aliases_used": file_aliases_used,
        }

    missing_from_all_files = sorted(expected_columns - all_mapped_columns)

    return {
        "schema_path": str(schema_path),
        "file_count": len(files),
        "expected_columns": sorted(expected_columns),
        "column_aliases": alias_map,
        "files": files_result,
        "aliases_used": {canonical: sorted(originals) for canonical, originals in sorted(aliases_used.items())},
        "extra_in_files": sorted(all_extra_columns),
        "missing_from_all_files": missing_from_all_files,
    }


def _print_main_report(result: dict[str, Any]) -> None:
    if "error" in result:
        print(f"ERROR: {result['error']}")
        return

    print("=" * 70)
    print("SCHEMA COMPARISON REPORT")
    print("=" * 70)
    print(f"Directory: {result['directory']}")
    print(f"Data root: {result['data_root']}")
    print(f"Files scanned: {result['file_count']}")
    print()

    print("FILES:")
    for filename, details in result["files"].items():
        print(f"  - {filename}: {details['column_count']} columns")
    print()

    print(
        f"UNION COLUMNS ({len(result['union_columns'])}): "
        + ", ".join(result["union_columns"])
    )
    print(
        f"INTERSECTION COLUMNS ({len(result['intersection_columns'])}): "
        + ", ".join(result["intersection_columns"])
    )

    if result["extra_columns"]:
        print()
        print("COLUMNS NOT PRESENT IN EVERY FILE:")
        for column, files_that_have_it in result["extra_columns"].items():
            print(f"  - {column}: present in {', '.join(files_that_have_it)}")
    else:
        print()
        print("All columns are present in every file.")

    print()
    print("POSSIBLE DUPLICATES (similar names):")
    if result["possible_variants"]:
        for names in result["possible_variants"]:
            print(f"  - {' vs '.join(names)}")
    else:
        print("  (none)")


def _print_baseline_report(result: dict[str, Any]) -> None:
    print()
    print("=" * 70)
    print("BASELINE COMPARISON")
    print("=" * 70)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        available = result.get("available_files")
        if available:
            print("Available files:")
            for filename in available:
                print(f"  - {filename}")
        return

    print(f"BASELINE: {result['baseline']} ({result['baseline_column_count']} columns)")
    print()
    if result["matches"]:
        print(f"MATCHES BASELINE ({len(result['matches'])} files):")
        print("  " + ", ".join(result["matches"]))
    else:
        print("MATCHES BASELINE: none")
    print()

    if result["differs"]:
        print("DIFFERS FROM BASELINE:")
        for filename, diff in result["differs"].items():
            print(f"  {filename}:")
            for column in diff["extra"]:
                print(f"    + {column} (extra)")
            for column in diff["missing"]:
                print(f"    - {column} (missing)")
    else:
        print("All files match the baseline.")


def _print_schema_check_report(result: dict[str, Any]) -> None:
    print()
    print("=" * 70)
    print("SCHEMA.YAML COMPARISON")
    print("=" * 70)

    if "error" in result:
        print(f"ERROR: {result['error']}")
        return

    print(f"Schema: {result['schema_path']}")
    print(f"Expected columns ({len(result['expected_columns'])}): {', '.join(result['expected_columns'])}")
    print()

    for filename, details in result["files"].items():
        print(f"{filename}:")
        print(f"  Columns in file: {details['column_count']}")
        if details["missing_from_schema_columns"]:
            print(f"  Missing from file: {', '.join(details['missing_from_schema_columns'])}")
        if details["extra_not_in_schema_columns"]:
            print(f"  Extra in file: {', '.join(details['extra_not_in_schema_columns'])}")
        if details["aliases_used"]:
            used = ", ".join(
                [f"{source} -> {target}" for source, target in sorted(details["aliases_used"].items())]
            )
            print(f"  Aliases used: {used}")
        if (
            not details["missing_from_schema_columns"]
            and not details["extra_not_in_schema_columns"]
            and not details["aliases_used"]
        ):
            print("  Matches schema exactly.")
        print()

    if result["aliases_used"]:
        print("ALIASES USED ACROSS FILES:")
        for canonical, originals in result["aliases_used"].items():
            print(f"  - {canonical}: {', '.join(originals)}")
        print()

    if result["extra_in_files"]:
        print(f"EXTRA COLUMNS IN FILES (not in schema): {', '.join(result['extra_in_files'])}")
    else:
        print("No extra columns found in files.")

    if result["missing_from_all_files"]:
        print(
            "MISSING FROM ALL FILES (schema expects these): "
            + ", ".join(result["missing_from_all_files"])
        )
    else:
        print("Every schema column appears in at least one file.")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare Excel schemas in a dataset raw/ directory.")
    parser.add_argument(
        "--data-root",
        type=Path,
        default=None,
        metavar="PATH",
        help="External data root (sets DATA_ROOT). Default: DATA_ROOT env or sibling ExcelIngestion_Data.",
    )
    parser.add_argument(
        "--dataset",
        default=DEFAULT_DATASET,
        metavar="NAME",
        help=f"Dataset name under DATA_ROOT/{{env}} (default: {DEFAULT_DATASET}).",
    )
    parser.add_argument(
        "--env",
        choices=("dev", "prod"),
        default=DEFAULT_ENV,
        metavar="ENV",
        help=f"Environment under DATA_ROOT (default: {DEFAULT_ENV}).",
    )
    parser.add_argument(
        "--baseline",
        default=None,
        metavar="FILENAME",
        help="Compare all files against one baseline Excel filename.",
    )
    parser.add_argument(
        "--check-against",
        action="store_true",
        help="Compare file headers against dataset config/schema.yaml (supports aliases).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        metavar="PATH",
        help="Write complete structured report JSON to this path.",
    )
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    if args.data_root is not None:
        os.environ["DATA_ROOT"] = str(args.data_root.expanduser().resolve())

    dataset_root = get_dataset_path(args.env, args.dataset)
    raw_dir = dataset_root / "raw"
    schema_path = dataset_root / "config" / "schema.yaml"

    result: dict[str, Any] = {
        "dataset": args.dataset,
        "env": args.env,
        "dataset_root": str(dataset_root),
    }

    main_report = compare_schemas(raw_dir)
    result["schema_comparison"] = main_report
    _print_main_report(main_report)

    exit_code = 1 if "error" in main_report else 0

    if args.baseline:
        baseline_report = compare_against_baseline(raw_dir, args.baseline)
        result["baseline_comparison"] = baseline_report
        _print_baseline_report(baseline_report)
        if "error" in baseline_report:
            exit_code = 1

    if args.check_against:
        schema_report = check_against_schema(raw_dir, schema_path)
        result["check_against_schema"] = schema_report
        _print_schema_check_report(schema_report)
        if "error" in schema_report:
            exit_code = 1

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        with args.output.open("w", encoding="utf-8") as handle:
            json.dump(result, handle, indent=2)
        print()
        print(f"Report written to: {args.output}")

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
